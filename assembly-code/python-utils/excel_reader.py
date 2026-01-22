import openpyxl

df = openpyxl.load_workbook("../PARM-table.xlsx")
df1 = df.active

# for row in range(3, df1.max_row):
#     for col in df1.iter_cols(1, df1.max_column):
#         value = col[row].value
#         if value is not None:
#             print(col[row].value)

dct={}

def reader(df1):
    instructions = []
    values = []

    # Read instructions
    for row in df1.iter_rows(4,51):
        value = row[4].value
        if value is not None:
            instructions.append(row[4].value)

    # Read opCode
    values.extend(read_bits(4, 6))
    values.extend(read_bits(8, 12))
    values.extend(read_bits(14, 29))
    values.extend(read_bits(31, 32))
    values.extend(read_bits(34, 35))
    values.extend(read_bits(37, 51))

    return instructions, values

def read_bits(begining_row_nb,end_row_nb):
    values = []
    for row in df1.iter_rows(begining_row_nb, end_row_nb):
        if row[10].value is not None:
            bits = []
        for col in range(10, 25):
            value = row[col].value
            if value is not None:
                bits.append(value)
        values.append(bits)
    return values